#-*- coding:utf-8 -*-
'''
Created on 2019. 12. 11.
Updated on 2022. 1. 6

@author: 100su_MOM
'''


# Python 3 program for Elo Rating 
import math 
from openpyxl import Workbook
from openpyxl import load_workbook
from pprint import pprint
import datetime
now = datetime.datetime.now()
nowDatetime = now.strftime('%Y%m%d%H%M%S')

# from operator import itemgetter, attrgetter
import gspread
from oauth2client.service_account import ServiceAccountCredentials
scope = [
'https://spreadsheets.google.com/feeds',
'https://www.googleapis.com/auth/drive',
]
json_file_name = 'hamicgamedata_key.json'
credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
gc = gspread.authorize(credentials)

write_data = Workbook()
write_ws = write_data.active
write_ws.title = 'GameData'
#write_ws = write_data.create_sheet('Players')
write_ws.append(['Date', 'Player 1', 'Race 1', 'Team 1', 'Tier 1', 'WL 1', 'Player 2', 'Race 2','Team 2', 'Tier 2',  'WL 2', 'MAP', 'League', 'T1 종변요청', 'T2 종변요청'])

#종족 변수
Race = ["T", "P", "Z", "t", "p", "z", "( T )", "( P )", "( Z )", "( t )", "( p )", "( z )", "(T)", "(P)", "(Z)", "( T)", "( P)", "( Z)", "(T )", "(P )", "(Z )", " T ", " P ", " Z ", 
        "(t)", "(p)", "(z)", "( t)", "( p)", "( z)", "(t )", "(p )", "(z )", " t ", " p ", " z ", " T", " P", " Z", "T ", "P ", "Z ", 
        "(테란)", "(토스)", "(저그)", "t승", "p승", "z승", "t패", "p패", "z패", "T승", "P승", "Z승", "T패", "P패", "Z패"]
   
gcMAP1 = gc.open("HAMIC_Player_ID").worksheet('MAPNAME')
totalMAP = []
totalMAPsearch = []
cntrow = -1
gcMAP2 = gcMAP1.get_all_values()
for idxR in range(len(gcMAP2)):
    row_data = gcMAP2[idxR]
    totalMAP.append([])
    cntrow += 1
    #print(row_data)
    for m in range(len(row_data)):  
        if row_data[m]: 
            totalMAP[cntrow].append(row_data[m])
        if row_data[m]:            
            totalMAPsearch.append(row_data[m])    
    print(totalMAP[cntrow])

print(totalMAPsearch)

#pprint(totalMAP, indent=4, width=150)

#종족 변수
RaceALL = ["T", "P", "Z", "t", "p", "z", "( T )", "( P )", "( Z )", "( t )", "( p )", "( z )", "(T)", "(P)", "(Z)", "( T)", "( P)", "( Z)", "(T )", "(P )", "(Z )", " T ", " P ", " Z ", 
        "(t)", "(p)", "(z)", "( t)", "( p)", "( z)", "(t )", "(p )", "(z )", " t ", " p ", " z ", " T", " P", " Z", "T ", "P ", "Z ", 
        "(테란)", "(토스)", "(저그)", "t승", "p승", "z승", "t패", "p패", "z패", "T승", "P승", "Z승", "T패", "P패", "Z패"]

TIERALL = ["TEAM", "HEL0", "HEL1", "HEL2", "AEL", "MEL", "IEL", "CEL1", "CEL2", "CEL3"]
pureTIER = ["HEL0", "HEL1", "HEL2", "AEL", "MEL", "IEL", "CEL1", "CEL2", "CEL3"]

TeamNames = ['STB', 'FW', 'AT', 'IDB', 'R2F', 'TSH','AV', 'LT', 'MOM', "F1", 'F1W', 'ITS', 'KaL', 'PrP', 'MYM', 'HBS', 'NF','PGS', 'FA']
Tier = ["HEL0", "HEL1", "HEL2", "AEL", "MEL", "IEL", "CEL1", "CEL2", "CEL3"]

#TIERALL = ["MELWAR"]


GameCounter = 0   
TEAMGameCounter = 0   

for idxTIER in TIERALL:
    try:
        f = open(idxTIER + ".txt", 'r', encoding='utf-8')
    except:
        continue
    
    lines = f.read().splitlines()
    
    if idxTIER == "TEAM":
        LEAGUE = "Team league"
        
        datachecked = 0
        for line in lines:
            print(line)

            #날짜 인식
            if '2020.' in line:
                idx = line.find('2020.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2019.' in line:
                idx = line.find('2019.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2021.' in line:
                idx = line.find('2021.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2022.' in line:
                idx = line.find('2022.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2023.' in line:
                idx = line.find('2023.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2024.' in line:
                idx = line.find('2024.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            elif '2025.' in line:
                idx = line.find('2025.')
                GameDate = line[idx:idx+10]
                datachecked = 1
            else:
                if not datachecked:
                    GameDate = '2020.x.x'
            
            for searchMAP in totalMAPsearch:
                idxMAP = line.lower().find(searchMAP.lower())
                if idxMAP > 0:
                    for totalMAProw in totalMAP:
                        if searchMAP in totalMAProw:
                            MAPinfo = totalMAProw[0]   
                    print(MAPinfo)
                    break
            
            #문자열 분할
            if ' vs ' not in line.lower() and '] - [' in line:
                idxTEAM = line.find('] - [')
                temTEAM1 = line[:idxTEAM-2]
                temTEAM2 = line[idxTEAM+7:]
                TEAM1 = temTEAM1.strip()
                TEAM2 = temTEAM2.strip()
            
            if ' vs ' in line.lower() and 'HAMIC' not in line:
                GameCounter += 1                
                idxDIV = line.lower().find(' vs ')
                lineP1 = line[:idxDIV+1]
                lineP2 = line[idxDIV+3:]
                            
                #등급 분할 
                for idxpureTIER in pureTIER:
                    if idxpureTIER in lineP1:
                        TIER1 = idxpureTIER
                        lineP1afterTIER = lineP1.replace(TIER1, "")
                    if idxpureTIER in lineP2:
                        TIER2 = idxpureTIER
                        lineP2afterTIER = lineP2.replace(TIER2, "")
                
                #승패 분할
                if " 승  " in lineP1afterTIER:
                    WL1 = "Win"
                    lineP1afterWL = lineP1afterTIER.replace(" 승 ", "")
                elif "승  " in lineP1afterTIER:
                    WL1 = "Win"
                    lineP1afterWL = lineP1afterTIER.replace("승 ", "")
                elif " 승 " in lineP1afterTIER:
                    WL1 = "Win"
                    lineP1afterWL = lineP1afterTIER.replace(" 승", "")
                elif " W  " in lineP1afterTIER:
                    WL1 = "Win"
                    lineP1afterWL = lineP1afterTIER.replace(" W ", "")
                elif " L  " in lineP1afterTIER:
                    WL1 = "Lose"
                    lineP1afterWL = lineP1afterTIER.replace(" L ", "")
                elif " 패" in lineP1afterTIER:
                    WL1 = "Lose"
                    lineP1afterWL = lineP1afterTIER.replace(" 패 ", "")
                elif "패 " in lineP1afterTIER:
                    WL1 = "Lose"
                    lineP1afterWL = lineP1afterTIER.replace("패  ", "")
                if " 승  " in lineP2afterTIER:
                    WL2 = "Win"
                    lineP2afterWL = lineP2afterTIER.replace(" 승 ", "")
                elif "승  " in lineP2afterTIER:
                    WL2 = "Win"
                    lineP2afterWL = lineP2afterTIER.replace("승 ", "")
                elif " 승 " in lineP2afterTIER:
                    WL2 = "Win"
                    lineP2afterWL = lineP2afterTIER.replace(" 승", "")
                elif " W " in lineP2afterTIER:
                    WL2 = "Win"
                    lineP2afterWL = lineP2afterTIER.replace(" W ", "")
                elif " L " in lineP2afterTIER:
                    WL2 = "Lose"
                    lineP2afterWL = lineP2afterTIER.replace(" L ", "")
                elif " 패" in lineP2afterTIER:
                    WL2 = "Lose"
                    lineP2afterWL = lineP2afterTIER.replace(" 패 ", "")
                elif "패 " in lineP2afterTIER:
                    WL2 = "Lose"
                    lineP2afterWL = lineP2afterTIER.replace("패  ", "")
                if "기권" in lineP1afterTIER or "기권패" in lineP1afterTIER:
                    WL1 = "LBD"
                    lineP1afterWL = lineP1afterTIER.replace("기권패", "")
                    WL2 = "WBD"
                elif "기권" in lineP2afterTIER or "기권패" in lineP2afterTIER:
                    WL2 = "LBD"
                    lineP2afterWL = lineP2afterTIER.replace("기권패", "")
                    WL1 = "WBD"
                                 
                #종족 분할
                cntRace = -1
                # 종족
                
                change1 = ' '
                change2 = ' '
                if ' -> ' in lineP1afterWL:
                    idxchange = lineP1afterWL.find(' -> ')
                    temline = lineP1afterWL[:idxchange-2] + lineP1afterWL[idxchange+4:]
                    lineP1afterWL = temline
                    change2 = 'O'
                if ' -> ' in lineP2afterWL:
                    idxchange = lineP2afterWL.find(' -> ')
                    temline = lineP2afterWL[:idxchange-2] + lineP2afterWL[idxchange+4:]
                    lineP2afterWL = temline    
                    change1 = 'O'
                    
                #ID 분할
                lineP1 = lineP1afterWL.strip()
                lineP2 = lineP2afterWL.strip()
                
                idxP1 = lineP1.find(" ")
                idxP2 = lineP2.find(" ")
                
                temRace1 = lineP1[idxP1:]
                temRace2 = lineP2[idxP2:]
                
                PlayerID1 = str(lineP1[:idxP1].strip())
                PlayerID2 = str(lineP2[:idxP2].strip())
                
                cntRace = -1
                for temRace in Race:
                    cntRace += 1
                    if temRace in temRace1:                       
                        break
                if cntRace%3 == 0:
                    Race1 = 'T'
                elif cntRace%3 == 1:
                    Race1 = 'P'
                else:
                    Race1 = 'Z'
                
                cntRace = -1                
                for temRace in Race:
                    cntRace += 1
                    if temRace in temRace2:                    
                        break
                if cntRace%3 == 0:
                    Race2 = 'T'
                elif cntRace%3 == 1:
                    Race2 = 'P'
                else:
                    Race2 = 'Z'
                 
                         
                ThisGameData = [GameDate, PlayerID1, Race1, TEAM1, TIER1, WL1, PlayerID2, Race2, TEAM2, TIER2, WL2, MAPinfo, LEAGUE, change1, change2]
                write_ws.append(ThisGameData)
    else:
        pre_line_all = []
        chk_game = 0
        chk_date = 0
        for line in lines:
            #날짜 인식
            if '2020.' in line:
                idx = line.find('2020.')
                GameDate = line[idx:idx+10]
                chk_date = 1
            if '2021.' in line:
                idx = line.find('2021.')
                GameDate = line[idx:idx+10]
                chk_date = 1
            if '2022.' in line:
                idx = line.find('2022.')
                GameDate = line[idx:idx+10]
                chk_date = 1
            if '2023.' in line:
                idx = line.find('2023.')
                GameDate = line[idx:idx+10]
                chk_date = 1
            if '2024.' in line:
                idx = line.find('2024.')
                GameDate = line[idx:idx+10]
                chk_date = 1
            if '2025.' in line:
                idx = line.find('2025.')
                GameDate = line[idx:idx+10]
                chk_date = 1
                    
            if '승' in line and '패' in line:
                pre_line_all.append(line)
                chk_game += 1
                    
            if chk_date == 1:
                
                for ii in range(chk_game):
                    pre_line = pre_line_all[ii]
                
                    #record
                    GameCounter += 1
                    if " : " in pre_line:
                        idxDIV = pre_line.find(' : ')
                        lineP1 = pre_line[:idxDIV]
                        lineP2 = pre_line[idxDIV+3:]
                    elif " vs " in pre_line.lower():
                        idxDIV = pre_line.lower().find(' vs ')
                        lineP1 = pre_line[:idxDIV]
                        lineP2 = pre_line[idxDIV+4:]
                    else:
                        idxDIV = pre_line.find(':')
                        lineP1 = pre_line[:idxDIV]
                        lineP2 = pre_line[idxDIV+1:]
                    
                    
                    if "승" in lineP1:
                        WL1 = "Win"
                        WL2 = "Lose"
                    else:
                        WL1 = "Lose"
                        WL2 = "Win"
                    
                    lineP1T = lineP1.split()     
                    lineP2T = lineP2.split() 
                    
                    print(pre_line)
                    print(lineP1T)
                    print(lineP2T)
                    
                    
                    PlayerID1 = lineP1T[0]
                    PlayerID2 = lineP2T[1]
                    
                    
                    temRace1 = lineP1T[1]
                    temRace2 = lineP2T[2]
                    
                    cntRace = -1
                    for temRace in RaceALL:
                        cntRace += 1
                        if temRace in temRace1:                       
                            break
                    if cntRace%3 == 0:
                        Race1 = 'T'
                    elif cntRace%3 == 1:
                        Race1 = 'P'
                    else:
                        Race1 = 'Z'
                        
                    cntRace = -1
                    for temRace in RaceALL:
                        cntRace += 1
                        if temRace in temRace2:                       
                            break
                    if cntRace%3 == 0:
                        Race2 = 'T'
                    elif cntRace%3 == 1:
                        Race2 = 'P'
                    else:
                        Race2 = 'Z'
        
                        
                    TEAM1 = 'xx'
                    TEAM2 = 'xx'
                    
                    TIER1 = idxTIER
                    TIER2 = idxTIER
                    LEAGUE = idxTIER + " league"
                        
                    MAPinfo = -1
                    for searchMAP in totalMAPsearch:
                        #print(searchMAP.upper())
                        lineupper = pre_line.upper()
                        #print(lineupper)
                        idxMAP = -1
                        idxMAP = lineupper.find(searchMAP.upper())
                        if idxMAP >= 0:
                            for totalMAProw in totalMAP:
                                tempmap = searchMAP.upper()
                                for idxM in range(len(totalMAProw)):
                                    if tempmap == totalMAProw[idxM].upper():
                                        MAPinfo = totalMAProw[0]  
                                        #lineAfterMAP = line.replace(searchMAP, "")  
                            print(MAPinfo)
                            break    
                        
                    if MAPinfo == -1:
                        MAPinfo = "not found"
        
                    
                    ThisGameData = [GameDate, PlayerID1, Race1, TEAM1, TIER1, WL1, PlayerID2, Race2, TEAM2, TIER2, WL2, MAPinfo, LEAGUE]
                    #print(lineP1)
                    #print(lineP2)
                    #print(ThisGameData)
                    write_ws.append(ThisGameData)
            
                #initialization
                chk_game = 0
                chk_date = 0
                pre_line_all = []
 
    f.close()


OutputFILE = "hamictotal_Gamedata_" + nowDatetime + ".xlsx"
write_data.save(OutputFILE)  

print(str(GameCounter) + ' Games Counted.')
print("Game Data Read from TXT to EXCEL -------")





#################
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

# Function to calculate the Probability 
def Change(nickname, Players):
    tempexistence = 0
    idxROW = 0
    idxCOL = 0
    for idxROW in range(len(Players)):
        for idxCOL in range(len(Players[idxROW])):
            if idxCOL == 0 or idxCOL == 1 or idxCOL == 3:
                continue
            else: 
                if Players[idxROW][idxCOL]: 
                    tempID = str(Players[idxROW][idxCOL])                
                    nickname = str(nickname.strip())         
                    if tempID.upper().split() == nickname.upper().split():
                        officialID = str(Players[idxROW][2]) 
                        officialID.split()       
                        TeamP = str(Players[idxROW][0])
                        tempexistence = 1
                        break
                    else:
                        #print("NOT included")
                        officialID = nickname
                        officialID.split()  
                        tempexistence = 0
                        TeamP = "XXX"
                else:
                        #print("NOT included")
                        officialID = nickname
                        officialID.split()  
                        tempexistence = 0
                        TeamP = "XXX"
        if tempexistence == 1: 
            break
    
    #print(str(Players[idxROW][idxCOL]) + tempID)
    #print(str(nickname) + "-->" + officialID + ":" + str(tempexistence))
    return [officialID, tempexistence, TeamP]

#################################
# 선수 리스트 dB 작성
#load_playerID = load_workbook("HAMIC_Player_ID.xlsx")

Players = []
cnt = -1
cntT = 0
cntP = 0
cntZ = 0
for a in enumerate(TeamNames):
    print(a[1])
    if a[1] == 'F1':
        a[1] == 'F1W'
        
    try:
        gdata_open = gc.open("HAMIC_Player_ID").worksheet((a[1]))
        gdata = gdata_open.get_all_values()
    except:
        continue
    
    #load_teamplayer = load_playerID[a[1]] 
    for idxR in range(len(gdata)):
        row_data = gdata[idxR]
        if row_data[0] in Tier:
            temp_row = [a[1]]
            for idxCOL in range(len(row_data)):
                if row_data[idxCOL]:
                    temp_row.append(row_data[idxCOL])
            
            
            Players.append(temp_row)
            cnt += 1
            print(Players[cnt])
            if row_data[2] == 'T':
                cntT += 1
            elif row_data[2] == 'P':
                cntP += 1
            else:
                cntZ += 1
                

print('Current total HAMIC players = ' + str(cnt+1))                        
print('Current total Terran players = ' + str(cntT))  
print('Current total Protoss players = ' + str(cntP))  
print('Current total Zerg players = ' + str(cntZ))


#################################

filenamebefore = OutputFILE
load_data = load_workbook(filenamebefore)
load_Game = load_data['GameData'] 

Year = ["2019", "2020", "2021", "2022", "2023", "2024", "2025"]

rowcnt = -1
for row in load_Game.rows:
    rowcnt += 1
    existence1 = 0
    existence2 = 0
    if row[0].value[0:4] in Year:
        # Player 1 ID check
        if row[1].value:
            nickname = str(row[1].value)
        else:
            nickname = "NotExist"
        [officialID, existence1, TeamP1] = Change(nickname, Players)
        row[1].value = str(officialID)
        row[3].value = TeamP1
        
        ft_RED = Font(color=colors.RED)
        if not existence1:            
            load_Game.cell(row = rowcnt+1, column = 2).font = ft_RED
        if TeamP1 == "XXX":
            load_Game.cell(row = rowcnt+1, column = 4).font = ft_RED
        # Player 2 ID check
        if row[6].value:
            nickname = str(row[6].value)
        else:
            nickname = "NotExist"
        [officialID, existence2, TeamP2] = Change(nickname, Players)
        row[6].value = str(officialID)
        row[8].value = TeamP2
        if not existence2:
            load_Game.cell(row = rowcnt+1, column = 7).font = ft_RED  
        if TeamP2 == "XXX":
            load_Game.cell(row = rowcnt+1, column = 9).font = ft_RED    
        
  
# 엑셀 파일 저장하기
filenameafter = filenamebefore
load_data.save(filenameafter)
print("Read game data Completed!")


from openpyxl.styles.colors import BLACK

 

#TIERALL = ["HEL1", "HEL2", "AEL", "MEL", "IEL", "CEL1", "CEL2"]

# Function to calculate the Probability 
def Probability(rating1, rating2):   
    return 1.0 * 1.0 / (1 + 1.0 * math.pow(10, 1.0 * (rating1 - rating2) / 400)) 

  
# Function to calculate Elo rating 
# K is a constant. 
# d determines whether 
# Player A wins or Player B.  
def EloRating(Ra, Rb, K, tierDiff, winner):   
    # To calculate the Winning 
    # Probability of Player B 
    Pb = Probability(Ra, Rb) 
    # To calculate the Winning 
    # Probability of Player A 
    Pa = Probability(Rb, Ra) 
    # Case -1 When Player A wins 
    # Updating the Elo Ratings 
    if tierDiff == 0:        
        if winner == 1: 
            Ra = Ra + K * (1 - Pa) 
            Rb = Rb + K * (0 - Pb)   
        # Case -2 When Player B wins 
        # Updating the Elo Ratings 
        elif winner == 2: 
            Ra = Ra + K * (0 - Pa) 
            Rb = Rb + K * (1 - Pb) 
        elif winner == 3:
            Ra = Ra + K/2
            Rb = Rb - K/2
        elif winner == 4:
            Ra = Ra - K/2
            Rb = Rb + K/2
    elif tierDiff < 0:  #P1 tier가 높음
        if winner == 1 or winner == 3:  #P1이 승리
            Ra += K/3
            Rb -= K/3
        else: # P2가 승리
            Ra -= K/3*2
            Rb += K/3*2
    elif tierDiff > 0:  #P2 tier가 높음
        if winner == 1 or winner == 3:  #P1이 승리
            Ra += K/3*2
            Rb -= K/3*2
        else: # P2가 승리
            Ra -= K/3
            Rb += K/3
            
    return [Ra, Rb]
    #print("Updated Ratings:-") 
    #print("Ra =", round(Ra, 6)," Rb =", round(Rb, 6)) 
  

#################################

filenamebefore = "ELO_MAP_Data.xlsx"
load_data = load_workbook(filenamebefore, data_only=True)


PlayersELO = []
rowcnt = 0
  
load_ELO = load_data["ELO"]        

for row in load_ELO.rows:    
    if row[1].value in pureTIER:
        rowcnt += 1
        temp = []
        for idx in range(10):
            if row[idx].value == None:
                temp.append(0)
            else:
                temp.append(row[idx].value)
            
        # TPZ 종족별 승패 6 , 전체 승패 2, 팀리그/개인리그 승패 4,
        for idx in range(12):
            temp.append(0)
            
        #  ELO 증가분 
        temp.append(row[7].value)  
        
        temp[7] = round(temp[7], 2)
        temp[2] = str(temp[2])
        
        [officialID, existence1, TeamP1] = Change(temp[2], Players)
        temp[2] = str(officialID)
        PlayersELO.append(temp)
            
            

print('총 ' + str(rowcnt) + '명의  선수 검색')  
#print(PlayersELO[10][10])     
print("Previsou ELO Load Completed!")

# Game data load
#filenameGame = "HAMIC_official_Game_Data_RAW.xlsx"
load_game = load_workbook(filenameafter, data_only=True)
try:
    load_games = load_game['GameData'] 
except:
    load_games = load_game.active


# ELO 파일 업데이트하기 
write_data = Workbook()
write_ws = write_data.active
write_ws.title = 'Updated Games'
#write_ws = write_data.create_sheet('Players')
#write_ws.append(['순위', '등급', '선수', '종족', '승', '패', '승률', 'ELO','우승', '준우승', ' +승 vs T', ' +패 vs T', ' +승 vs P', ' +패 vs P', ' +승 vs Z', ' +패 vs Z', ' +승 ', ' +패', ' +팀리그승', ' +팀리그패', ' +개인리그승', ' +개인리그패', 'ELO 증가분'])



for row in load_games.rows:
    addedROW = []
    for idxR in range(len(row)):
        addedROW.append(row[idxR].value)
    write_ws.append(addedROW)
    #print(str(row[1].value) + ' vs ' + str(row[6].value))
    if row[0].value[0:4] == "2019" or row[0].value[0:4] == "2020" or row[0].value[0:4] == "2021" or row[0].value[0:4] == "2022" or row[0].value[0:4] == "2023" or row[0].value[0:4] == "2024" or row[0].value[0:4] == "2025":
        # 선수 찾기 
        foundcheck = 0
        idxP1 = -1
        idxP2 = -1
        for idxP in range(rowcnt):
            if str(row[1].value).upper().strip() == str(PlayersELO[idxP][2]).upper().strip():
                idxP1 = idxP
                break
        
        for idxP in range(rowcnt):
            if str(row[6].value).upper().strip() == str(PlayersELO[idxP][2]).upper().strip():
                idxP2 = idxP
                break
        
        if idxP1 == -1:
            idxP1 = rowcnt
            rowcnt += 1
            temp = []
            temp.append(0)
            temp.append(str(row[4].value))
            temp.append(str(row[1].value))
            temp.append(str(row[2].value))
            temp.append(0)
            temp.append(0)
            temp.append(0)
                
            if str(row[4].value) == "HEL1":
                tempELO = 1600.00
            elif str(row[4].value) == "HEL2":
                tempELO = 1550.00
            elif str(row[4].value) == "AEL":
                tempELO = 1500.00
            elif str(row[4].value) == "MEL":
                tempELO = 1450.00
            elif str(row[4].value) == "IEL":
                tempELO = 1400.00
            elif str(row[4].value) == "CEL1":
                tempELO = 1350.00
            elif str(row[4].value) == "CEL2":
                tempELO = 1300.00   
            elif str(row[4].value) == "CEL3":
                tempELO = 1300.00  
            elif str(row[4].value) == "HEL0":
                tempELO = 1600.00   
            
            temp.append(tempELO)
            for idx in range(14):
                temp.append(0)   
            temp.append(round(tempELO, 2))
            PlayersELO.append(temp)
            
        if idxP2 == -1:
            idxP2 = rowcnt
            rowcnt += 1
            temp = []
            temp.append(0)
            temp.append(str(row[9].value))
            temp.append(str(row[6].value))
            temp.append(str(row[7].value))
            
            temp.append(0)
            temp.append(0)
            temp.append(0)
                
            if str(row[9].value) == "HEL1":
                tempELO = 1600
            elif str(row[9].value) == "HEL2":
                tempELO = 1550
            elif str(row[9].value) == "AEL":
                tempELO = 1500
            elif str(row[9].value) == "MEL":
                tempELO = 1450
            elif str(row[9].value) == "IEL":
                tempELO = 1400
            elif str(row[9].value) == "CEL1":
                tempELO = 1350
            elif str(row[9].value) == "CEL2":
                tempELO = 1300   
            elif str(row[9].value) == "CEL3":
                tempELO = 1300.00  
            elif str(row[9].value) == "HEL0":
                tempELO = 1600.00   
             
            
            temp.append(tempELO)
            for idx in range(14):
                temp.append(0)  
            temp.append(round(tempELO, 2))
            PlayersELO.append(temp)

        # P1 선수 전적 업데이트
        if row[5].value == "Win":
            if row[7].value == "T":
                PlayersELO[idxP1][10] += 1
            elif row[7].value == "P":
                PlayersELO[idxP1][12] += 1
            elif row[7].value == "Z":
                PlayersELO[idxP1][14] += 1
            
            PlayersELO[idxP1][4] += 1
            PlayersELO[idxP1][16] += 1
            
            if row[12].value == "Team league":
                PlayersELO[idxP1][18] += 1
            else:
                PlayersELO[idxP1][20] += 1
        else:
            if row[7].value == "T":
                PlayersELO[idxP1][11] += 1
            elif row[7].value == "P":
                PlayersELO[idxP1][13] += 1
            elif row[7].value == "Z":
                PlayersELO[idxP1][15] += 1
            
            PlayersELO[idxP1][5] += 1
            PlayersELO[idxP1][17] += 1
            
            if row[12].value == "Team league":
                PlayersELO[idxP1][19] += 1
            else:
                PlayersELO[idxP1][21] += 1
        
        # P2 선수 전적 업데이트
        if row[10].value == "Win":
            if row[2].value == "T":
                PlayersELO[idxP2][10] += 1
            elif row[2].value == "P":
                PlayersELO[idxP2][12] += 1
            elif row[2].value == "Z":
                PlayersELO[idxP2][14] += 1
            
            PlayersELO[idxP2][4] += 1
            PlayersELO[idxP2][16] += 1
            
            if row[12].value == "Team league":
                PlayersELO[idxP2][18] += 1
            else:
                PlayersELO[idxP2][20] += 1
        else:
            if row[2].value == "T":
                PlayersELO[idxP2][11] += 1
            elif row[2].value == "P":
                PlayersELO[idxP2][13] += 1
            elif row[2].value == "Z":
                PlayersELO[idxP2][15] += 1
            
            PlayersELO[idxP2][5] += 1
            PlayersELO[idxP2][17] += 1
            
            if row[12].value == "Team league":
                PlayersELO[idxP2][19] += 1
            else:
                PlayersELO[idxP2][21] += 1
                
        # ELO update
        temp1 = PlayersELO[idxP1][7]
        temp2 = PlayersELO[idxP2][7]
        # tier=-1,0,1 (P1 tier - P2 tier: P1티어가 높으면 -1, 낮으면 1 같으면 0, tier2, nogame (부전승)
        # d = 1, P1 승리, d = 0, P2승리
        cnttierP1 = 0
        cnttierP2 = 0
        for temptier in TIERALL:
            if row[4].value == temptier:
                break
            else:
                cnttierP1 += 1
                
        for temptier in TIERALL:
            if row[9].value == temptier:
                break
            else:
                cnttierP2 += 1
        tierDiffer = cnttierP1 - cnttierP2
        if row[12].value == "Team league":
            Kpoint = 60;
        else:
            Kpoint = 30;
        
        if tierDiffer == 0:
            if row[5].value == "Win":
                WINP = 1
                [temp1, temp2] = EloRating(temp1, temp2, Kpoint, tierDiffer, WINP) 
            elif row[5].value == "Lose":
                WINP = 2
                [temp1, temp2] = EloRating(temp1, temp2, Kpoint, tierDiffer, WINP) 
            else:
                [temp1, temp2] = [temp1, temp2]             
        else:                    
            [temp1, temp2] = [temp1, temp2] 
        

        PlayersELO[idxP1][7] = round(temp1, 2)
        PlayersELO[idxP2][7] = round(temp2, 2)
        
 
# ELO 증가분 파악하기
for idxP in range(rowcnt):
    PlayersELO[idxP][22] = round(PlayersELO[idxP][7] - PlayersELO[idxP][22], 2)
     

from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
    
    
write_wsall = write_data.create_sheet("ELO") 
write_wsall.append(['순위', '등급', '선수', '종족', '승', '패', '승률', 'ELO','우승', '준우승', ' +승 vs T', ' +패 vs T', ' +승 vs P', ' +패 vs P', ' +승 vs Z', ' +패 vs Z', ' +승 ', ' +패', ' +팀리그승', ' +팀리그패', ' +개인리그승', ' +개인리그패', 'ELO 증가분'])

# 티어별 쓰기
for temptier in TIERALL:
    write_wstemp = write_data.create_sheet(temptier) 
    write_wstemp.append(['순위', '등급', '선수', '종족', '승', '패', '승률', 'ELO','우승', '준우승', ' +승 vs T', ' +패 vs T', ' +승 vs P', ' +패 vs P', ' +승 vs Z', ' +패 vs Z', ' +승 ', ' +패', ' +팀리그승', ' +팀리그패', ' +개인리그승', ' +개인리그패', 'ELO 증가분'])

    for idxP in range(rowcnt):
        if PlayersELO[idxP][1] in pureTIER:
            write_wsall.append(PlayersELO[idxP])
        if PlayersELO[idxP][1] == temptier:
            write_wstemp.append(PlayersELO[idxP])
            

TIERALL = ["ELO", "HEL0", "HEL1", "HEL2", "AEL", "MEL", "IEL", "CEL1", "CEL2", "CEL3"]
for temptier in TIERALL:
    rowcnt = -1
    write_wstemp = write_data[temptier] 

    for row in write_wstemp.rows:
        rowcnt += 1
        if row[3].value == "T":
            ft_COL = Font(color="00B0F0")
        elif row[3].value == "Z":
            ft_COL = Font(color=colors.RED)    
        elif row[3].value == "P":
            ft_COL = Font(color=colors.YELLOW)   
    
        if row[3].value == "T" or row[3].value == "P" or row[3].value == "Z":
            write_wstemp.cell(row = rowcnt+1, column = 2).font = ft_COL
            write_wstemp.cell(row = rowcnt+1, column = 3).font = ft_COL
            write_wstemp.cell(row = rowcnt+1, column = 4).font = ft_COL
            write_wstemp.cell(row = rowcnt+1, column = 2).fill = PatternFill(fgColor=BLACK, fill_type = "solid") 
            write_wstemp.cell(row = rowcnt+1, column = 3).fill = PatternFill(fgColor=BLACK, fill_type = "solid") 
            write_wstemp.cell(row = rowcnt+1, column = 4).fill = PatternFill(fgColor=BLACK, fill_type = "solid") 
   
           
OutputFILE = "ELO_MAP_Data_updatedIN" + nowDatetime + ".xlsx"
write_data.save(OutputFILE)  

print("ALL_Completed!")
